"""Builds evals/promptfoo/ranking/data/eval_dataset.json from the
Job_Matching_Golden_Dataset xlsx (README / Job Listings / Ranking (Ground
Truth) / Candidates sheets) — replacing the hand-written synthetic
placeholder dataset with a real, documented golden set.

Usage:
    python3 build_ranking_dataset_from_xlsx.py <path-to-xlsx> [--split dev|test|all]

Notes on fit with the ranking eval's metrics (see also README update):
  - Each candidate has exactly 10 jobs, not the 15 the original synthetic
    set used. recall@10/precision@10 over a 10-item pool are DEGENERATE
    (top-10 of a 10-item list is the whole list, so those two metrics
    would be constant regardless of ranking quality) — this script does
    NOT carry those fields forward. Use precision@5 / recall@5 (both
    equal here, since ground_truth_top5 always has exactly 5 'Y' rows)
    and ndcg@10 instead, plus Spearman rank correlation, which the
    dataset's own README explicitly recommends.
  - ndcg is computed against the continuous `ground_truth_rule_score`
    (0-100), not a bucketed 0-3 label — that's strictly more information
    and NDCG is scale-invariant, so there's no reason to throw resolution
    away by bucketing into ground_truth_label.
"""
import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl --break-system-packages")


def _split_list(s: str) -> list[str]:
    return [p.strip() for p in (s or "").split(";") if p.strip()]


def _job_text(row: dict) -> str:
    exp = f"{row['experience_required_min']}-{row['experience_required_max']} years"
    return (
        f"{row['job_title']} at {row['company']}. {row['job_category']} role. "
        f"Location: {row['job_location']} ({row['work_mode']}). "
        f"Experience required: {exp}. "
        f"Required skills: {row['required_skills']}. "
        f"{row['job_description']}"
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("--split", choices=["dev", "test", "all"], default="all",
                         help="Which eval_split to include (Candidates sheet). "
                              "Use 'dev' while iterating on weights, 'test' for a final check, "
                              "'all' (default) to use every candidate.")
    parser.add_argument("-o", "--output", default=None,
                         help="Output path (default: evals/promptfoo/ranking/data/eval_dataset.json)")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    # ---- Candidates sheet ----
    ws = wb["Candidates"]
    rows = list(ws.iter_rows(values_only=True))
    header, body = rows[0], rows[1:]
    idx = {name: i for i, name in enumerate(header)}
    candidates = {}
    for r in body:
        split = r[idx["eval_split"]]
        if args.split != "all" and split != args.split:
            continue
        cid = r[idx["candidate_id"]]
        candidates[cid] = {
            "candidate_json": {
                "name": r[idx["name"]],
                "category": r[idx["category"]],
                "experience_years": r[idx["years_experience_est"]],
                "location": r[idx["location"]],
                "skills": _split_list(r[idx["top_skills"]]),
                "target_roles": _split_list(r[idx["job_titles"]]),
                "domain": r[idx["category"]],
                "eval_split": split,
            },
            "jobs": [],
        }
    print(f"Loaded {len(candidates)} candidates (split={args.split})")

    # ---- Job Listings sheet ----
    ws = wb["Job Listings"]
    rows = list(ws.iter_rows(values_only=True))
    header, body = rows[0], rows[1:]
    idx = {name: i for i, name in enumerate(header)}
    jobs_by_id = {}
    for r in body:
        row = {name: r[i] for name, i in idx.items()}
        jobs_by_id[row["job_id"]] = row

    # ---- Ranking (Ground Truth) sheet ----
    ws = wb["Ranking (Ground Truth)"]
    rows = list(ws.iter_rows(values_only=True))
    header, body = rows[0], rows[1:]
    idx = {name: i for i, name in enumerate(header)}
    n_attached = 0
    for r in body:
        row = {name: r[i] for name, i in idx.items()}
        cid = row["candidate_id"]
        if cid not in candidates:
            continue
        job = jobs_by_id.get(row["job_id"])
        if not job:
            continue
        candidates[cid]["jobs"].append({
            "job_id": row["job_id"],
            "job_text": _job_text(job),
            "ground_truth_rule_score": float(row["ground_truth_rule_score"]),
            "ground_truth_rank": int(row["ground_truth_rank"]),
            "ground_truth_top5": row["ground_truth_top5"] == "Y",
            "ground_truth_label": row["ground_truth_label"],
        })
        n_attached += 1

    empty = [cid for cid, c in candidates.items() if not c["jobs"]]
    for cid in empty:
        del candidates[cid]
    if empty:
        print(f"Dropped {len(empty)} candidates with no matched jobs: {empty[:5]}{'...' if len(empty) > 5 else ''}")

    print(f"Attached {n_attached} job rows across {len(candidates)} candidates")

    out_path = Path(args.output) if args.output else (
        Path(__file__).resolve().parents[1] / "promptfoo" / "ranking" / "data" / "eval_dataset.json"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps({"candidates": candidates}, indent=2))
    print(f"Wrote {out_path}")

    # Also emit the promptfoo tests/candidates.yaml listing every candidate_id
    tests_path = out_path.parent.parent / "tests" / "candidates.yaml"
    tests_lines = [f'- vars: {{ candidate_id: "{cid}" }}' for cid in candidates]
    tests_path.write_text(
        "# Auto-generated by build_ranking_dataset_from_xlsx.py — do not hand-edit.\n"
        "# Re-run that script (with --split as needed) to regenerate.\n"
        + "\n".join(tests_lines) + "\n"
    )
    print(f"Wrote {tests_path} ({len(candidates)} test cases)")


if __name__ == "__main__":
    main()
