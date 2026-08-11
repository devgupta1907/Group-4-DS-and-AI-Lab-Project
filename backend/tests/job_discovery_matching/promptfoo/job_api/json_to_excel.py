"""
Convert promptfoo's results.json into an Excel report.

Usage:
    python json_to_excel.py
    python json_to_excel.py path/to/results.json path/to/output.xlsx
"""

import json
import sys

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def load_results(results_path: str) -> list[dict]:
    with open(results_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    results = data.get("results", {}).get("results", [])
    if not results:
        results = data.get("results", [])
    return results


def build_rows(results: list[dict]) -> list[dict]:
    rows = []
    for r in results:
        success = r.get("success")
        if success is None:
            success = (r.get("gradingResult") or {}).get("pass")

        test_vars = r.get("vars", {})
        description = r.get("testCase", {}).get("description") or r.get(
            "description", "(no description)"
        )

        error_msg = r.get("error")
        if not error_msg:
            grading = r.get("gradingResult") or {}
            error_msg = grading.get("reason")

        response = r.get("response")
        output = response.get("output") if isinstance(response, dict) else r.get("output")
        output_str = json.dumps(output, ensure_ascii=False) if output is not None else ""

        rows.append(
            {
                "Description": description,
                "Query": test_vars.get("query", ""),
                "Location": test_vars.get("location", ""),
                "Country": test_vars.get("country", ""),
                "Limit": test_vars.get("limit", ""),
                "Result": "PASS" if success else "FAIL",
                "Error": error_msg or "",
                "Output (truncated)": output_str[:500],
            }
        )
    return rows


def write_excel(rows: list[dict], output_path: str) -> None:
    df = pd.DataFrame(rows)
    df.to_excel(output_path, index=False, sheet_name="Results")

    # Formatting pass with openpyxl
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    ws = wb["Results"]

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    body_font = Font(name="Arial")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    result_col_idx = list(df.columns).index("Result") + 1

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        result_cell = ws.cell(row=row_idx, column=result_col_idx)
        result_cell.fill = pass_fill if result_cell.value == "PASS" else fail_fill

    widths = {
        "Description": 32,
        "Query": 22,
        "Location": 14,
        "Country": 10,
        "Limit": 8,
        "Result": 10,
        "Error": 40,
        "Output (truncated)": 60,
    }
    for i, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col_name, 20)

    ws.freeze_panes = "A2"
    wb.save(output_path)


if __name__ == "__main__":
    in_path = sys.argv[1] if len(sys.argv) > 1 else "results.json"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "results.xlsx"

    results = load_results(in_path)
    rows = build_rows(results)
    write_excel(rows, out_path)

    passed = sum(1 for r in rows if r["Result"] == "PASS")
    print(f"Wrote {len(rows)} rows to {out_path} ({passed} passed, {len(rows) - passed} failed)")